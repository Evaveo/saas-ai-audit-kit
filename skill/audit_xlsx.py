"""
Génère un fichier Excel d'audit SaaS+IA, vide ou pré-rempli avec les résultats
fournis par les sous-agents d'audit.

USAGE :
    # Template vide
    python audit_xlsx.py --output audit_template.xlsx

    # Avec résultats
    python audit_xlsx.py --results audit_results.json --output AUDIT_filled.xlsx

    # Ne lancer que certains thèmes (gagne tokens + temps)
    python audit_xlsx.py --scope 02_Securite_Auth,10_Legal --output partial.xlsx

    # Filtrer selon le stade du produit (skip checks prématurés)
    python audit_xlsx.py --profile early-stage --output AUDIT_mvp.xlsx

    # Output machine-readable pour CI/CD
    python audit_xlsx.py --results audit_results.json --output AUDIT.xlsx --summary-json summary.json

NOUVEAUX statuts (v2) :
    - OK       : critère respecté
    - PARTIAL  : partiellement respecté (compte pour 0.5 OK dans le score)
    - KO       : critère non respecté
    - TODO     : prévu mais pas encore implémenté (n'est pas compté dans le score)
    - N/A      : ne s'applique pas

SCORE PONDÉRÉ (v2) : 🔴 compte 3×, 🟡 compte 2×, 🟢 compte 1×.

PROFILS (v2) :
    - all (défaut) : tous les checks
    - early-stage  : skip les checks tagués "profile":"launch+" ou "scale-only"
    - launch-ready : skip "scale-only"
    - scale        : tous les checks (identique à "all")

Le fichier `themes.json` doit être dans le même dossier que ce script.
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


SCRIPT_DIR = Path(__file__).parent

# Poids par criticité — utilisé pour le score pondéré
WEIGHTS = {"🔴": 3, "🟡": 2, "🟢": 1}

# Profils — quels tags d'items skipper
PROFILE_SKIPS = {
    "all":          set(),
    "early-stage":  {"launch+", "scale-only"},
    "launch-ready": {"scale-only"},
    "scale":        set(),
}

VALID_STATUSES = ("OK", "PARTIAL", "KO", "TODO", "N/A")

# Lignes / plages fixes des feuilles thème (formules Score / Synthèse)
THEME_SCORE_ROW = 3
THEME_HEADER_ROW = 4
THEME_DATA_START_ROW = 5
THEME_DATA_LAST_ROW = 200


# ─── Formules Excel (scores par thème) ───────────────────────────────────────

def _crit_weight_array(crit_rng: str) -> str:
    return f'IF({crit_rng}="🔴",3,IF({crit_rng}="🟡",2,IF({crit_rng}="🟢",1,0)))'


def _status_points_array(score_rng: str) -> str:
    return f'IF({score_rng}="OK",1,IF({score_rng}="PARTIAL",0.5,IF({score_rng}="KO",0,0)))'


def _evaluated_mask(score_rng: str) -> str:
    return f'(({score_rng}="OK")+({score_rng}="PARTIAL")+({score_rng}="KO"))'


def raw_score_formula(score_rng: str) -> str:
    """Score brut : (OK + PARTIAL/2) / (OK + PARTIAL + KO)."""
    denom = f'COUNTIF({score_rng},"OK")+COUNTIF({score_rng},"PARTIAL")+COUNTIF({score_rng},"KO")'
    num = f'COUNTIF({score_rng},"OK")+COUNTIF({score_rng},"PARTIAL")*0.5'
    return f'=IF({denom}=0,"",({num})/({denom}))'


def theme_local_score_ranges() -> tuple[str, str]:
    """Plages Score / Crit. sur la feuille courante (sans préfixe d'onglet)."""
    score_rng = f"$G${THEME_DATA_START_ROW}:$G${THEME_DATA_LAST_ROW}"
    crit_rng = f"$E${THEME_DATA_START_ROW}:$E${THEME_DATA_LAST_ROW}"
    return score_rng, crit_rng


def weighted_num_formula(score_rng: str, crit_rng: str) -> str:
    return f'=SUMPRODUCT({_crit_weight_array(crit_rng)}*{_status_points_array(score_rng)})'


def weighted_den_formula(score_rng: str, crit_rng: str) -> str:
    w = _crit_weight_array(crit_rng)
    return f'=SUMPRODUCT({w}*{_evaluated_mask(score_rng)})'


def weighted_score_formula(num_cell: str, den_cell: str) -> str:
    return f'=IF({den_cell}=0,"",{num_cell}/{den_cell})'


def theme_score_ranges(sheet_name: str) -> tuple[str, str]:
    """Plages Score (G) et Crit. (E) pour les formules d'un onglet thème."""
    prefix = f"'{sheet_name}'!"
    score_rng = f"{prefix}$G${THEME_DATA_START_ROW}:$G${THEME_DATA_LAST_ROW}"
    crit_rng = f"{prefix}$E${THEME_DATA_START_ROW}:$E${THEME_DATA_LAST_ROW}"
    return score_rng, crit_rng


# ─── Styling ─────────────────────────────────────────────────────────────────

BORDER_THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color="0F172A")
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="64748B")
META_FONT = Font(name="Calibri", size=10, color="475569")
CELL_FONT = Font(name="Calibri", size=11, color="0F172A")
SECTION_HEADER_FONT = Font(name="Calibri", size=12, bold=True, color="0F172A")
SCORE_LABEL_FONT = Font(name="Calibri", size=11, bold=True, color="475569")
SCORE_VALUE_FONT = Font(name="Calibri", size=11, bold=True, color="0F172A")

CRIT_FILLS = {
    "🔴": PatternFill("solid", fgColor="FEE2E2"),
    "🟡": PatternFill("solid", fgColor="FEF3C7"),
    "🟢": PatternFill("solid", fgColor="DCFCE7"),
}

STATUS_FILLS = {
    "OK":      PatternFill("solid", fgColor="BBF7D0"),
    "PARTIAL": PatternFill("solid", fgColor="FED7AA"),  # orange clair
    "KO":      PatternFill("solid", fgColor="FECACA"),
    "TODO":    PatternFill("solid", fgColor="DBEAFE"),  # bleu clair
    "N/A":     PatternFill("solid", fgColor="E2E8F0"),
}


def style_header_row(ws, row: int, columns: int) -> None:
    for col in range(1, columns + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN


# ─── Filtering helpers ────────────────────────────────────────────────────────

def filter_themes(themes: list[dict], scope: list[str] | None, profile: str) -> list[dict]:
    """Filtre thèmes selon --scope et items selon --profile."""
    skip_tags = PROFILE_SKIPS.get(profile, set())

    filtered = []
    for theme in themes:
        if scope and theme["id"] not in scope:
            continue
        # Filter items by profile tag
        kept_items = []
        for item in theme.get("items", []):
            item_profile = item.get("profile", "all")
            if item_profile in skip_tags:
                continue
            kept_items.append(item)
        if kept_items:
            t = dict(theme)
            t["items"] = kept_items
            filtered.append(t)
    return filtered


# ─── Build ───────────────────────────────────────────────────────────────────

def build_workbook(themes: list[dict], results: dict | None = None,
                   audited_repo: str | None = None, audit_date: str | None = None,
                   profile: str = "all") -> Workbook:
    """Construit le classeur. Si `results` est fourni, pré-remplit Score + Commentaire."""
    wb = Workbook()

    # ── Mode d'emploi ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Mode d'emploi"

    ws["A1"] = "Grille d'audit SaaS + IA"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    subtitle = "Audit automatisé via skill Claude — repo : " + (audited_repo or "n/a")
    if audit_date:
        subtitle += f" — date : {audit_date}"
    if profile and profile != "all":
        subtitle += f" — profil : {profile}"
    ws["A2"] = subtitle if results else "Cadre réutilisable pour auditer n'importe quel SaaS qui intègre de l'IA."
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:E2")

    sections = [
        (None, None),
        ("Méthode", None),
        (None, f"• Chaque onglet = un thème de l'audit ({len(themes)} thèmes inclus dans ce rapport)."),
        (None, "• Score : OK / PARTIAL / KO / TODO / N/A (dropdown sur la colonne Score)."),
        (None, "• PARTIAL compte pour 0.5 OK. TODO et N/A sont exclus du score."),
        (None, "• Score pondéré : 🔴 compte 3×, 🟡 compte 2×, 🟢 compte 1×."),
        (None, "• Chaque onglet thème affiche en ligne 3 le score brut et le score pondéré (formules)."),
        (None, "• La feuille 'Synthèse' récupère ces scores depuis chaque onglet thème."),
        (None, None),
        ("Légende statut", None),
        ("OK", "Critère respecté"),
        ("PARTIAL", "Partiellement respecté (compte 0.5)"),
        ("KO", "Critère non respecté"),
        ("TODO", "Prévu, pas encore implémenté (non compté)"),
        ("N/A", "Ne s'applique pas"),
        (None, None),
        ("Légende criticité", None),
        ("🔴", "Bloquant launch — fixer avant ouverture publique (poids 3×)"),
        ("🟡", "Important — à corriger sous 30 jours (poids 2×)"),
        ("🟢", "Amélioration continue — nice to have (poids 1×)"),
        (None, None),
        ("Seuils de décision", None),
        ("> 90%", "Ready to scale — focus growth"),
        ("80-90%", "Ready to launch — fix les 🟡 sous 30j"),
        ("65-80%", "Ready bêta privée — fix tous les 🔴"),
        ("< 65%", "Pas prêt pour la prod"),
        (None, None),
        ("Ordre d'attaque", None),
        ("1.", "Tous les 🔴 d'abord (bloquants)"),
        ("2.", "Légal + Données (risque pénal / régulateur)"),
        ("3.", "Sécurité (risque incident)"),
        ("4.", "IA Coûts + Gouvernance (risque facture / image)"),
        ("5.", "Le reste : amélioration continue"),
    ]
    row = 3
    for label, val in sections:
        if label and not val:
            ws.cell(row=row, column=1, value=label).font = SECTION_HEADER_FONT
        elif label:
            ws.cell(row=row, column=1, value=label).font = META_FONT
            ws.cell(row=row, column=2, value=val).font = META_FONT
        elif val:
            ws.cell(row=row, column=2, value=val).font = META_FONT
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 70

    # ── Synthèse (placeholder, filled at the end) ────────────────────────────
    synthese_ws = wb.create_sheet("Synthèse")

    # ── Theme sheets ─────────────────────────────────────────────────────────
    status_validation = DataValidation(
        type="list", formula1='"OK,PARTIAL,KO,TODO,N/A"', allow_blank=True, showDropDown=False,
    )
    status_validation.error = "Valeur attendue : OK, PARTIAL, KO, TODO ou N/A"
    status_validation.errorTitle = "Statut invalide"

    for theme in themes:
        ws = wb.create_sheet(theme["id"])
        theme_results = (results.get("themes", {}).get(theme["id"], {})) if results else {}

        ws["A1"] = theme["name"]
        ws["A1"].font = TITLE_FONT
        ws.merge_cells("A1:I1")

        ws["A2"] = theme["description"]
        ws["A2"].font = SUBTITLE_FONT
        ws.merge_cells("A2:I2")

        score_rng, crit_rng = theme_local_score_ranges()
        sr = THEME_SCORE_ROW

        ws.cell(row=sr, column=1, value="Score brut").font = SCORE_LABEL_FONT
        raw_cell = ws.cell(row=sr, column=2, value=raw_score_formula(score_rng))
        raw_cell.number_format = "0%"
        raw_cell.font = SCORE_VALUE_FONT
        raw_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=sr, column=3, value="Score pondéré").font = SCORE_LABEL_FONT
        num_ref = f"$K${sr}"
        den_ref = f"$L${sr}"
        ws.cell(row=sr, column=11, value=weighted_num_formula(score_rng, crit_rng))
        ws.cell(row=sr, column=12, value=weighted_den_formula(score_rng, crit_rng))
        wgt_cell = ws.cell(row=sr, column=4, value=weighted_score_formula(num_ref, den_ref))
        wgt_cell.number_format = "0%"
        wgt_cell.font = SCORE_VALUE_FONT
        wgt_cell.alignment = Alignment(horizontal="center", vertical="center")

        for col in (1, 2, 3, 4):
            ws.cell(row=sr, column=col).border = BORDER_THIN
        for score_col in ("B", "D"):
            ws.conditional_formatting.add(
                f"{score_col}{sr}",
                CellIsRule(operator="greaterThanOrEqual", formula=["0.9"], fill=PatternFill("solid", fgColor="BBF7D0")),
            )
            ws.conditional_formatting.add(
                f"{score_col}{sr}",
                CellIsRule(operator="between", formula=["0.65", "0.8999"], fill=PatternFill("solid", fgColor="FEF3C7")),
            )
            ws.conditional_formatting.add(
                f"{score_col}{sr}",
                CellIsRule(operator="lessThan", formula=["0.65"], fill=PatternFill("solid", fgColor="FECACA")),
            )

        headers = [
            "#", "Question Chef de Projet", "Critère Développeur", "Comment tester",
            "Crit.", "Auto", "Score", "Commentaire", "Outils & IA",
        ]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=THEME_HEADER_ROW, column=col, value=h)
        style_header_row(ws, THEME_HEADER_ROW, len(headers))

        for i, item in enumerate(theme["items"], start=THEME_DATA_START_ROW):
            ws.cell(row=i, column=1, value=item["id"]).font = CELL_FONT
            ws.cell(row=i, column=2, value=item.get("pm_question", "")).font = CELL_FONT
            ws.cell(row=i, column=3, value=item["label"]).font = CELL_FONT
            ws.cell(row=i, column=4, value=item["how"]).font = CELL_FONT
            crit_cell = ws.cell(row=i, column=5, value=item["crit"])
            crit_cell.font = CELL_FONT
            crit_cell.fill = CRIT_FILLS[item["crit"]]
            crit_cell.alignment = Alignment(horizontal="center", vertical="center")

            auto_cell = ws.cell(row=i, column=6, value=item.get("automation", ""))
            auto_cell.font = CELL_FONT
            auto_cell.alignment = Alignment(horizontal="center", vertical="center")

            item_result = theme_results.get(item["id"], {})
            status_value = item_result.get("status", "")
            status_cell = ws.cell(row=i, column=7, value=status_value)
            status_cell.alignment = Alignment(horizontal="center", vertical="center")
            status_cell.font = Font(name="Calibri", size=11, bold=True)
            if status_value in STATUS_FILLS:
                status_cell.fill = STATUS_FILLS[status_value]

            notes = item_result.get("notes", "") if item_result else ""
            evidence = item_result.get("evidence", "") if item_result else ""
            caveat = item.get("manual_caveat") or ""
            parts = []
            if notes:
                parts.append(notes)
            if evidence and evidence != "null":
                parts.append(f"📍 {evidence}")
            if caveat:
                parts.append(f"⚠️ {caveat}")
            ws.cell(row=i, column=8, value="\n".join(parts)).font = CELL_FONT

            tools = item.get("recommended_tools") or ""
            ai_role = item.get("ai_role") or ""
            tools_parts = []
            if tools:
                tools_parts.append(f"🔧 {tools}")
            if ai_role:
                tools_parts.append(f"🤖 {ai_role}")
            ws.cell(row=i, column=9, value="\n".join(tools_parts)).font = Font(name="Calibri", size=9, color="475569")

            for col in range(1, 10):
                cell = ws.cell(row=i, column=col)
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal or "left",
                    vertical="top",
                    wrap_text=True,
                )
                cell.border = BORDER_THIN

        last_row = THEME_HEADER_ROW + len(theme["items"])
        status_validation.add(f"G{THEME_DATA_START_ROW}:G{last_row}")

        for status, fill_color in [("OK", "BBF7D0"), ("PARTIAL", "FED7AA"), ("KO", "FECACA"), ("TODO", "DBEAFE"), ("N/A", "E2E8F0")]:
            ws.conditional_formatting.add(
                f"G{THEME_DATA_START_ROW}:G{last_row}",
                FormulaRule(
                    formula=[f'G{THEME_DATA_START_ROW}="{status}"'],
                    fill=PatternFill("solid", fgColor=fill_color),
                ),
            )

        widths = [6, 40, 42, 38, 8, 8, 12, 38, 42]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.column_dimensions["K"].hidden = True
        ws.column_dimensions["L"].hidden = True
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[THEME_HEADER_ROW].height = 28
        ws.row_dimensions[THEME_SCORE_ROW].height = 22
        for r in range(THEME_DATA_START_ROW, last_row + 1):
            ws.row_dimensions[r].height = 70

        ws.freeze_panes = f"C{THEME_DATA_START_ROW}"
        ws.add_data_validation(status_validation)

    # ── Synthèse ─────────────────────────────────────────────────────────────
    ws = synthese_ws
    ws["A1"] = "Synthèse de l'audit"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:L1")

    subtitle = "Score brut = OK / (OK + KO). Score pondéré = somme(poids × statut) — 🔴=3, 🟡=2, 🟢=1. PARTIAL=0.5. TODO/N/A exclus."
    if audited_repo:
        subtitle = f"Repo : {audited_repo}" + (f"  •  Date : {audit_date}" if audit_date else "")
    ws["A2"] = subtitle
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:L2")

    headers = ["#", "Thème", "🔴", "🟡", "🟢", "OK", "PART.", "KO", "TODO", "N/A", "Score %", "Score pond. %"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    row = 5
    for i, theme in enumerate(themes, start=1):
        sheet_name = theme["id"]
        score_rng, crit_rng = theme_score_ranges(sheet_name)
        sheet_ref = f"'{sheet_name}'"
        ws.cell(row=row, column=1, value=i).font = CELL_FONT
        ws.cell(row=row, column=2, value=theme["name"]).font = CELL_FONT
        ws.cell(row=row, column=3, value=f'=COUNTIF({crit_rng},"🔴")').font = CELL_FONT
        ws.cell(row=row, column=4, value=f'=COUNTIF({crit_rng},"🟡")').font = CELL_FONT
        ws.cell(row=row, column=5, value=f'=COUNTIF({crit_rng},"🟢")').font = CELL_FONT
        ws.cell(row=row, column=6, value=f'=COUNTIF({score_rng},"OK")').font = CELL_FONT
        ws.cell(row=row, column=7, value=f'=COUNTIF({score_rng},"PARTIAL")').font = CELL_FONT
        ws.cell(row=row, column=8, value=f'=COUNTIF({score_rng},"KO")').font = CELL_FONT
        ws.cell(row=row, column=9, value=f'=COUNTIF({score_rng},"TODO")').font = CELL_FONT
        ws.cell(row=row, column=10, value=f'=COUNTIF({score_rng},"N/A")').font = CELL_FONT
        score_cell = ws.cell(row=row, column=11, value=f"={sheet_ref}!$B${THEME_SCORE_ROW}")
        score_cell.number_format = "0%"
        score_cell.font = Font(name="Calibri", size=11, bold=True)
        wgt_cell = ws.cell(row=row, column=12, value=f"={sheet_ref}!$D${THEME_SCORE_ROW}")
        wgt_cell.number_format = "0%"
        wgt_cell.font = Font(name="Calibri", size=11, bold=True)

        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(
                horizontal="center" if col != 2 else "left",
                vertical="center",
                wrap_text=True,
            )
            cell.border = BORDER_THIN
            if col == 3:
                cell.fill = CRIT_FILLS["🔴"]
            elif col == 4:
                cell.fill = CRIT_FILLS["🟡"]
            elif col == 5:
                cell.fill = CRIT_FILLS["🟢"]

        row += 1

    total_row = row
    ws.cell(row=total_row, column=2, value="TOTAL").font = Font(name="Calibri", size=11, bold=True)
    for col_letter, col_idx in [("C", 3), ("D", 4), ("E", 5), ("F", 6), ("G", 7), ("H", 8), ("I", 9), ("J", 10)]:
        ws.cell(row=total_row, column=col_idx,
                value=f"=SUM({col_letter}5:{col_letter}{total_row-1})").font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=total_row, column=11,
            value=f'=IF((F{total_row}+G{total_row}+H{total_row})=0,"",(F{total_row}+G{total_row}*0.5)/(F{total_row}+G{total_row}+H{total_row}))').number_format = "0%"
    ws.cell(row=total_row, column=11).font = Font(name="Calibri", size=12, bold=True, color="0F172A")
    wgt_num_refs = ",".join(f"'{t['id']}'!$K${THEME_SCORE_ROW}" for t in themes)
    wgt_den_refs = ",".join(f"'{t['id']}'!$L${THEME_SCORE_ROW}" for t in themes)
    global_wgt_cell = ws.cell(
        row=total_row,
        column=12,
        value=f"=IF(SUM({wgt_den_refs})=0,\"\",SUM({wgt_num_refs})/SUM({wgt_den_refs}))",
    )
    global_wgt_cell.number_format = "0%"
    global_wgt_cell.font = Font(name="Calibri", size=12, bold=True, color="0F172A")
    for col in range(1, 13):
        ws.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor="F1F5F9")
        ws.cell(row=total_row, column=col).border = BORDER_THIN

    for col_letter in ("K", "L"):
        ws.conditional_formatting.add(
            f"{col_letter}5:{col_letter}{total_row}",
            CellIsRule(operator="greaterThanOrEqual", formula=["0.9"], fill=PatternFill("solid", fgColor="BBF7D0")),
        )
        ws.conditional_formatting.add(
            f"{col_letter}5:{col_letter}{total_row}",
            CellIsRule(operator="between", formula=["0.65", "0.8999"], fill=PatternFill("solid", fgColor="FEF3C7")),
        )
        ws.conditional_formatting.add(
            f"{col_letter}5:{col_letter}{total_row}",
            CellIsRule(operator="lessThan", formula=["0.65"], fill=PatternFill("solid", fgColor="FECACA")),
        )

    widths = [6, 40, 8, 8, 8, 8, 8, 8, 8, 8, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[4].height = 24

    legend_row = total_row + 3
    ws.cell(row=legend_row, column=2, value="Seuils").font = Font(name="Calibri", size=11, bold=True)
    legends = [
        ("> 90%", "Ready to scale — focus growth", "BBF7D0"),
        ("80-90%", "Ready to launch — fix 🟡 sous 30j", "FEF3C7"),
        ("65-80%", "Ready bêta — fix tous les 🔴", "FEF3C7"),
        ("< 65%", "Pas prêt pour la prod", "FECACA"),
    ]
    for i, (k, v, color) in enumerate(legends, start=1):
        r = legend_row + i
        ws.cell(row=r, column=2, value=k).fill = PatternFill("solid", fgColor=color)
        ws.cell(row=r, column=2).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=v).font = META_FONT
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=12)

    ws.freeze_panes = "A5"

    current_idx = wb.sheetnames.index("Synthèse")
    if current_idx != 1:
        wb.move_sheet("Synthèse", offset=1 - current_idx)

    return wb


# ─── Score helpers ────────────────────────────────────────────────────────────

def _status_to_value(status: str) -> float | None:
    """OK→1, PARTIAL→0.5, KO→0, autres (TODO/N/A) → None (exclus du score)."""
    return {"OK": 1.0, "PARTIAL": 0.5, "KO": 0.0}.get(status)


def compute_weighted_scores(themes: list[dict], results: dict) -> dict[str, float]:
    """Score pondéré par thème — Σ(poids × valeur_statut) / Σ(poids des items évalués)."""
    out = {}
    items_by_id = {t["id"]: t.get("items", []) for t in themes}
    for theme_id, item_results in (results.get("themes", {}) or {}).items():
        items = {it["id"]: it for it in items_by_id.get(theme_id, [])}
        num, denom = 0.0, 0.0
        for iid, r in item_results.items():
            item = items.get(iid)
            if not item:
                continue
            value = _status_to_value(r.get("status", ""))
            if value is None:
                continue
            w = WEIGHTS.get(item.get("crit", ""), 1)
            num += w * value
            denom += w
        if denom > 0:
            out[theme_id] = num / denom
    return out


def compute_global_weighted(themes: list[dict], results: dict) -> float | None:
    """Score global pondéré — sur tous les items."""
    items_by_id = {t["id"]: t.get("items", []) for t in themes}
    num, denom = 0.0, 0.0
    for theme_id, item_results in (results.get("themes", {}) or {}).items():
        items = {it["id"]: it for it in items_by_id.get(theme_id, [])}
        for iid, r in item_results.items():
            item = items.get(iid)
            if not item:
                continue
            value = _status_to_value(r.get("status", ""))
            if value is None:
                continue
            w = WEIGHTS.get(item.get("crit", ""), 1)
            num += w * value
            denom += w
    return (num / denom) if denom > 0 else None


def compute_summary(themes: list[dict], results: dict) -> dict:
    """Synthèse JSON exhaustive — utile pour CI/CD ou dashboard."""
    status_counts = {s: 0 for s in VALID_STATUSES}
    by_theme = {}
    blockers = []  # 🔴 KO seulement
    todos = []     # Tout TODO
    items_by_id = {t["id"]: t for t in themes}

    for theme_id, item_results in (results.get("themes", {}) or {}).items():
        theme = items_by_id.get(theme_id)
        if not theme:
            continue
        theme_counts = {s: 0 for s in VALID_STATUSES}
        items_map = {it["id"]: it for it in theme.get("items", [])}
        for iid, r in item_results.items():
            status = r.get("status", "")
            if status in status_counts:
                status_counts[status] += 1
                theme_counts[status] += 1
            item = items_map.get(iid)
            if not item:
                continue
            if status == "KO" and item.get("crit") == "🔴":
                blockers.append({
                    "theme": theme_id, "id": iid, "label": item.get("label"),
                    "evidence": r.get("evidence"), "notes": r.get("notes"),
                })
            if status == "TODO":
                todos.append({
                    "theme": theme_id, "id": iid, "label": item.get("label"),
                    "crit": item.get("crit"), "notes": r.get("notes"),
                })
        by_theme[theme_id] = {
            "name": theme["name"],
            "counts": theme_counts,
            "raw_score": _raw_score(theme_counts),
            "weighted_score": compute_weighted_scores(themes, results).get(theme_id),
        }

    return {
        "audited_repo": results.get("audited_repo"),
        "audit_date": results.get("audit_date"),
        "total_items": sum(status_counts.values()),
        "status_counts": status_counts,
        "raw_score": _raw_score(status_counts),
        "weighted_score": compute_global_weighted(themes, results),
        "by_theme": by_theme,
        "blockers": blockers,
        "todos": todos,
    }


def _raw_score(counts: dict[str, int]) -> float | None:
    """OK + PARTIAL/2 sur (OK + PARTIAL + KO)."""
    ok = counts.get("OK", 0)
    partial = counts.get("PARTIAL", 0)
    ko = counts.get("KO", 0)
    denom = ok + partial + ko
    return ((ok + partial * 0.5) / denom) if denom else None


# ─── Entrypoint ──────────────────────────────────────────────────────────────

def main() -> int:
    parser = argparse.ArgumentParser(description="Build/fill SaaS+AI audit workbook")
    parser.add_argument("--results", type=str, default=None,
                        help="Path to audit_results.json (optional)")
    parser.add_argument("--output", type=str, required=True, help="Output .xlsx path")
    parser.add_argument("--themes", type=str, default=str(SCRIPT_DIR / "themes.json"),
                        help="Path to themes.json (default: alongside this script)")
    parser.add_argument("--scope", type=str, default=None,
                        help="Comma-separated theme IDs (e.g. '02_Securite_Auth,10_Legal')")
    parser.add_argument("--profile", type=str, default="all",
                        choices=list(PROFILE_SKIPS.keys()),
                        help="Filter items by maturity profile (default: all)")
    parser.add_argument("--summary-json", type=str, default=None,
                        help="Also write a machine-readable summary JSON (for CI/CD)")
    args = parser.parse_args()

    themes_path = Path(args.themes)
    if not themes_path.exists():
        print(f"ERROR: themes.json not found at {themes_path}", file=sys.stderr)
        return 1

    with themes_path.open("r", encoding="utf-8") as f:
        all_themes = json.load(f)["themes"]

    # Apply filters
    scope = [s.strip() for s in args.scope.split(",")] if args.scope else None
    themes = filter_themes(all_themes, scope, args.profile)

    if not themes:
        print(f"ERROR: no themes matched scope={args.scope} profile={args.profile}", file=sys.stderr)
        return 1

    results = None
    audited_repo = None
    audit_date = None
    if args.results:
        results_path = Path(args.results)
        if not results_path.exists():
            print(f"ERROR: results file not found at {results_path}", file=sys.stderr)
            return 1
        with results_path.open("r", encoding="utf-8") as f:
            results = json.load(f)
        audited_repo = results.get("audited_repo")
        audit_date = results.get("audit_date")

    wb = build_workbook(themes, results=results, audited_repo=audited_repo,
                        audit_date=audit_date, profile=args.profile)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"OK: saved to {output_path}")

    if results:
        summary = compute_summary(themes, results)
        sc = summary["status_counts"]
        raw = summary["raw_score"]
        wgt = summary["weighted_score"]
        raw_s = f"{raw:.0%}" if raw is not None else "n/a"
        wgt_s = f"{wgt:.0%}" if wgt is not None else "n/a"
        print(f"Summary: OK={sc['OK']} PARTIAL={sc['PARTIAL']} KO={sc['KO']} TODO={sc['TODO']} N/A={sc['N/A']}")
        print(f"  Score brut: {raw_s}  |  Score pondéré: {wgt_s}")
        print(f"  Blockers (KO): {len(summary['blockers'])}  |  TODO: {len(summary['todos'])}")

        if args.summary_json:
            summary_path = Path(args.summary_json)
            summary_path.parent.mkdir(parents=True, exist_ok=True)
            with summary_path.open("w", encoding="utf-8") as f:
                json.dump(summary, f, ensure_ascii=False, indent=2)
            print(f"  summary.json -> {summary_path}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
